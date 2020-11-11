
# coding: utf-8

# In[1]:


#!/usr/bin/python


# In[2]:


# packages to load 
# Check the versions of libraries
# Python version
#import win32com.client
#import win32com
import warnings
warnings.filterwarnings('ignore')
import sys
print('Python: {}'.format(sys.version))
sys.path.append('/usr/local/bin/python2')
import os
os.getcwd()
import pandas as pd
print(pd.__version__)
import glob
import datetime
import time
import shutil
#from openpyxl import load_workbook
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import (
    LineChart,
    Reference,
)
from openpyxl.chart.axis import DateAxis
import smtplib
import matplotlib.pyplot as plt
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formatdate
from email import encoders




# In[3]:


#Config_File = glob.glob("C:\\Users\erairoy\\Desktop\\Automation\\MTN\\CONFIG\\CONFIG_PYTHON.xlsx")
Config_File = glob.glob("/home/guineabisau/PM-Automation/MTN/CONFIG/CONFIG_PYTHON_UNIX.xlsx")
Configsheet_to_df_map = pd.read_excel(Config_File[0], sheet_name=None)
print('Configuration Loaded..')


# In[4]:
def Is_Input_File_Copied(opco,df_Config):
    try:
        
        Input_File_Prefix = df_Config[df_Config['CODENAME']=='INPUT_FILE_PREFIX']['CODEVALUE'].iloc[0]
        print(Input_File_Prefix)
        Input_File_Path = glob.glob(df_Config[df_Config['CODENAME']=='INPUT_FILE_PATH']['CODEVALUE'].iloc[0])
        print(Input_File_Path[0])
        Input_file_Path_to_Copy = glob.glob(df_Config[df_Config['CODENAME']=='EXTERNAL_FILE_PATH']['CODEVALUE'].iloc[0])
        print(Input_file_Path_to_Copy[0])
        File_name = Input_File_Prefix+datetime.datetime.today().strftime("%Y-%m-%d")+".xlsx"
        print(File_name)
        Input_file = shutil.copy2(Input_file_Path_to_Copy[0]+File_name, Input_File_Path[0]) # complete target filename given

        return "Success"
    except IndexError:
        print('Input file missing in External directory:'+Input_file_to_Copy)
        return "None"

def InputFile_Exists(opco,df_Config):
    try:
        
        Input_File_Path = df_Config[df_Config['CODENAME']=='INPUT_FILE_PATH']['CODEVALUE']
        Input_file = glob.glob(df_Config[df_Config['CODENAME']=='INPUT_FILE_PATH']['CODEVALUE'].iloc[0]+"MTN*.xlsx")
        return Input_file[0]
    except IndexError:
        print('Input file missing in the directory:'+Input_File_Path)
        return "None"


# In[5]:


#Copy range of cells as a nested list
#Takes: start cell, end cell, and sheet you want to copy from.
def CopyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
 
    return rangeSelected


# In[6]:


#Paste range
#Paste data from copyRange into template sheet
def PasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):
            
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1


# In[7]:


def Find_Stat(df_stat,entity):
    isExists ='N'
    #day = pd.to_datetime('today').strftime('%m/%d/%Y')
    #df_stat['Day'] = pd.Series(pd.to_datetime('today').strftime('%m/%d/%Y'), index=df_stat.index)
    #print(df_stat)
    try:
        df_row_stat = df_stat.loc[df_stat['Node Type']==entity,['Total number of backups','Backup failed','Backup Successful']]
        isExists ='Y'
        return isExists,df_row_stat.set_index('Total number of backups').T
    except IndexError:
        print('Failed to Fetch Individual Stat..')
        return isExists,df_stat
    
    


# In[8]:


def Enrich_Information(opco,entity,default_value,inputsheet_to_df_map):
    print(inputsheet_to_df_map)
    if opco == 'BIS':
        if entity == 'EMA':
            if inputsheet_to_df_map['PROCLOG'] and inputsheet_to_df_map['SOGCONFIG'] =='Done':
                return 'Done'
            else:
                return 'Not Done'
    if opco == 'SW':
        if entity == 'NGCRS':
            if inputsheet_to_df_map['APPFS'] =='Not Done':
                return default_value
            else:
                return 'N/A'
        elif entity == 'VS':
            if inputsheet_to_df_map['ORA_ARCHIVE'] =='Not Done':
                return default_value
            else:
                return 'N/A'    


# In[9]:


def Load_Format_Write(opco,Configsheet_to_df_map):
    #Load Input and Format Template Data
    df_Config = Configsheet_to_df_map[opco]
    Copied_File = Is_Input_File_Copied(opco,df_Config)
    if Copied_File == 'None':
        return
    Input_file = InputFile_Exists(opco,df_Config)
    
    if Input_file != "None":
        inputsheet_to_df_map = pd.read_excel(Input_file, sheet_name=None,skiprows=6)
        Template_File_Path = df_Config[df_Config['CODENAME']=='OUTPUT_TEMPLATE_PATH']['CODEVALUE'].iloc[0]
        Template_File_Name = df_Config[df_Config['CODENAME']=='TEMPLATE_FILE_NAME']['CODEVALUE'].iloc[0]
        Template_File = glob.glob(Template_File_Path+Template_File_Name)
        print(Input_file)
        print(Template_File[0])
        #print(inputsheet_to_df_map['AIR'])
        template_to_df_map = pd.read_excel(Template_File[0], sheet_name=None)
        #print(template_to_df_map['AIR'])
    else:
        print("No Input File is found")
        return
        
    #Manipulation on Non SDP Data    
    for i,row_entity in df_Config.loc[df_Config['CODENAME']=='ENTITY',['CODEVALUE']].iterrows():
        entity = row_entity['CODEVALUE']
        print(entity)
                        
        #print(template_to_df_map[entity])
        if entity not in ['SDP','Summary']:
            for index,row in template_to_df_map[entity].iterrows():
                try:
                    
                    #print(df_Config.loc[df_Config['CODENAME']=='MAP_'+entity,['CODEVALUE','CODEVALUE1']].iloc[0])
                    for i,row_source_dest in df_Config.loc[df_Config['CODENAME']=='MAP_'+entity,['CODEVALUE','CODEVALUE1','CODEVALUE2','CODEVALUE3']].iterrows():
                        #print(row_source_dest['CODEVALUE2'])
                        #print(row_source_dest['CODEVALUE3'])
                        if row_source_dest['CODEVALUE2']== 'DIRECT':
                            template_to_df_map[entity].loc[template_to_df_map[entity]['Node IP']==row['Node IP'],[row_source_dest['CODEVALUE1']]]=inputsheet_to_df_map[entity][inputsheet_to_df_map[entity]['Node-IP']==row['Node IP']].iloc[0][row_source_dest['CODEVALUE']]
                            #print(inputsheet_to_df_map[entity][inputsheet_to_df_map[entity]['Node-IP']==row['Node IP']].iloc[0][row_source_dest['CODEVALUE']])
                        elif row_source_dest['CODEVALUE2']== 'INDIRECT':
                            #print('INDIRECT:'+opco + " "+entity+" "+row_source_dest['CODEVALUE3'])
                            #print(inputsheet_to_df_map[entity][inputsheet_to_df_map[entity]['Node-IP']==row['Node IP']].iloc[0])
                            template_to_df_map[entity].loc[template_to_df_map[entity]['Node IP']==row['Node IP'],[row_source_dest['CODEVALUE1']]]=Enrich_Information(opco,entity,row_source_dest['CODEVALUE3'],inputsheet_to_df_map[entity][inputsheet_to_df_map[entity]['Node-IP']==row['Node IP']].iloc[0])
                            
                except IndexError:
                    continue
                #print(template_to_df_map[entity])
                #print('Formatting complete for :'+entity)    
        else:
            print("NON SDP formatting Complete..")
    
    #Manipulation on SDP Data
    entity='SDP'
    for index,row in template_to_df_map[entity].iterrows():
        try:
            for i,row_source_dest in df_Config.loc[df_Config['CODENAME']=='MAP_A_'+entity,['CODEVALUE','CODEVALUE1']].iterrows():
                template_to_df_map[entity].loc[template_to_df_map[entity]['Node A-IP']==row['Node A-IP'],[row_source_dest['CODEVALUE1']]]=inputsheet_to_df_map[entity][inputsheet_to_df_map[entity]['Node-IP']==row['Node A-IP']].iloc[0][row_source_dest['CODEVALUE']]
        except IndexError:
            continue
        try:                                       
            for i,row_source_dest in df_Config.loc[df_Config['CODENAME']=='MAP_B_'+entity,['CODEVALUE','CODEVALUE1']].iterrows():
                template_to_df_map[entity].loc[template_to_df_map[entity]['Node A-IP']==row['Node A-IP'],[row_source_dest['CODEVALUE1']]]=inputsheet_to_df_map[entity][inputsheet_to_df_map[entity]['Node-IP']==row['Node B-IP']].iloc[0][row_source_dest['CODEVALUE']]
        except IndexError:
            continue
        #print('Formatting complete for :'+entity)
    
    if opco == 'NG':
        entity='Summary'
        performance =0
        for index,row in template_to_df_map[entity].iterrows():
            count =0
            success = 0
            Falure = 0
            try:
                if row['Node Type'] not in ['Total','Performance']:
                    if row['Node Type'] == 'SDP': 
                        print("***************************************************")
                        #print(template_to_df_map[row['Node Type']].loc[template_to_df_map[row['Node Type']]['Node-A']=='Done',['Node-A']].count().iloc[0])
                        print(row['Node Type'])
                        print("***************************************************")
                        count = template_to_df_map[row['Node Type']]['Node A-IP'].count()+template_to_df_map[row['Node Type']]['Node B-IP'].count()
                        success = template_to_df_map[row['Node Type']].loc[template_to_df_map[row['Node Type']]['Node-A']=='Done',['Node-A']].count().iloc[0]+template_to_df_map[row['Node Type']].loc[template_to_df_map[row['Node Type']]['Node-B']=='Done',['Node-B']].count().iloc[0]
                        Falure = count - success
                        print('count:'+str(count) +' success:'+str(success)+' Falure:'+str(Falure))
                        #print(template_to_df_map[entity].loc[template_to_df_map[entity]['Node Type']==row['Node Type'],['Total number of backups']].iloc[0])
                        template_to_df_map[entity].loc[template_to_df_map[entity]['Node Type']==row['Node Type'],['Total number of backups']]=count
                        template_to_df_map[entity].loc[template_to_df_map[entity]['Node Type']==row['Node Type'],['Backup Successful']]=success
                        template_to_df_map[entity].loc[template_to_df_map[entity]['Node Type']==row['Node Type'],['Backup failed']]=Falure
                        #print(template_to_df_map[entity])
                    elif row['Node Type'] == 'AIR': 
                        print("***************************************************")
                        print(row['Node Type'])
                        print("***************************************************")
                        count = template_to_df_map[row['Node Type']]['TAPE Backup Status'].count()
                        success = template_to_df_map[row['Node Type']].loc[template_to_df_map[row['Node Type']]['TAPE Backup Status']=='Done',['TAPE Backup Status']].count().iloc[0]
                        Falure = count - success
                        print('count:'+str(count) +' success:'+str(success)+' Falure:'+str(Falure))
                        #print(template_to_df_map[entity].loc[template_to_df_map[entity]['Node Type']==row['Node Type'],['Total number of backups']].iloc[0])
                        template_to_df_map[entity].loc[template_to_df_map[entity]['Node Type']==row['Node Type'],['Total number of backups']]=count
                        template_to_df_map[entity].loc[template_to_df_map[entity]['Node Type']==row['Node Type'],['Backup Successful']]=success
                        template_to_df_map[entity].loc[template_to_df_map[entity]['Node Type']==row['Node Type'],['Backup failed']]=Falure
                    elif row['Node Type'] == 'CCN': 
                        print("***************************************************")
                        print(row['Node Type'])
                        print("***************************************************")
                        count = template_to_df_map[row['Node Type']]['Daily DBN backup'].count()
                        success = template_to_df_map[row['Node Type']].loc[template_to_df_map[row['Node Type']]['Daily DBN backup']=='Done',['Daily DBN backup']].count().iloc[0]
                        Falure = count - success
                        print('count:'+str(count) +' success:'+str(success)+' Falure:'+str(Falure))
                        #print(template_to_df_map[entity].loc[template_to_df_map[entity]['Node Type']==row['Node Type'],['Total number of backups']].iloc[0])
                        template_to_df_map[entity].loc[template_to_df_map[entity]['Node Type']==row['Node Type'],['Total number of backups']]=count
                        template_to_df_map[entity].loc[template_to_df_map[entity]['Node Type']==row['Node Type'],['Backup Successful']]=success
                        template_to_df_map[entity].loc[template_to_df_map[entity]['Node Type']==row['Node Type'],['Backup failed']]=Falure
                     
                    elif row['Node Type'] == 'OCC': 
                        print("***************************************************")
                        print(row['Node Type'])
                        print("***************************************************")
                        count = template_to_df_map[row['Node Type']]['FS Backup'].count()
                        success = template_to_df_map[row['Node Type']].loc[template_to_df_map[row['Node Type']]['FS Backup']=='Done',['FS Backup']].count().iloc[0]
                        Falure = count - success
                        print('count:'+str(count) +' success:'+str(success)+' Falure:'+str(Falure))
                        #print(template_to_df_map[entity].loc[template_to_df_map[entity]['Node Type']==row['Node Type'],['Total number of backups']].iloc[0])
                        template_to_df_map[entity].loc[template_to_df_map[entity]['Node Type']==row['Node Type'],['Total number of backups']]=count
                        template_to_df_map[entity].loc[template_to_df_map[entity]['Node Type']==row['Node Type'],['Backup Successful']]=success
                        template_to_df_map[entity].loc[template_to_df_map[entity]['Node Type']==row['Node Type'],['Backup failed']]=Falure
                    
                    elif row['Node Type'] == 'NGVS': 
                        print("***************************************************")
                        print(row['Node Type'])
                        print("***************************************************")
                        count = template_to_df_map[row['Node Type']]['DB backup'].count()+template_to_df_map[row['Node Type']]['FS Backup'].count()
                        success = template_to_df_map[row['Node Type']].loc[template_to_df_map[row['Node Type']]['DB backup']=='Done',['DB backup']].count().iloc[0]+template_to_df_map[row['Node Type']].loc[template_to_df_map[row['Node Type']]['FS Backup']=='Done',['FS Backup']].count().iloc[0]
                        Falure = count - success
                        print('count:'+str(count) +' success:'+str(success)+' Falure:'+str(Falure))
                        #print(template_to_df_map[entity].loc[template_to_df_map[entity]['Node Type']==row['Node Type'],['Total number of backups']].iloc[0])
                        template_to_df_map[entity].loc[template_to_df_map[entity]['Node Type']==row['Node Type'],['Total number of backups']]=count
                        template_to_df_map[entity].loc[template_to_df_map[entity]['Node Type']==row['Node Type'],['Backup Successful']]=success
                        template_to_df_map[entity].loc[template_to_df_map[entity]['Node Type']==row['Node Type'],['Backup failed']]=Falure
                        
                    else:
                        print('Node Not found..')
                    
                elif row['Node Type'] == 'Total':
                    print("***************************************************")
                    print(row['Node Type'])
                    #print(template_to_df_map['Summary']['Total number of backups'].sum())
                    print("***************************************************")
                    #template_to_df_map['Summary'].at['Total', 'Total number of backups'] = template_to_df_map['Summary']['Total number of backups'].sum()
                    count = template_to_df_map['Summary']['Total number of backups'].sum()
                    success = template_to_df_map['Summary']['Backup Successful'].sum()
                    #success = template_to_df_map[row['Node Type']].loc[template_to_df_map[row['Node Type']]['FS Backup']=='Done',['FS Backup']].count().iloc[0]
                    Falure = count - success
                    #print('count:'+str(count) )
                    #print(template_to_df_map[entity].loc[template_to_df_map[entity]['Node Type']==row['Node Type'],['Total number of backups']].iloc[0])
                    template_to_df_map[entity].loc[template_to_df_map[entity]['Node Type']==row['Node Type'],['Total number of backups']]=count
                    template_to_df_map[entity].loc[template_to_df_map[entity]['Node Type']==row['Node Type'],['Backup Successful']]=success
                    template_to_df_map[entity].loc[template_to_df_map[entity]['Node Type']==row['Node Type'],['Backup failed']]=Falure
                    print(template_to_df_map['Summary']['Total number of backups'].sum())    
                    performance=(success/count)

                    #print('Total')
                elif row['Node Type'] == 'Performance':
                    template_to_df_map[entity].loc[template_to_df_map[entity]['Node Type']==row['Node Type'],['Backup Successful']]=performance #str(performance)+'%'
                else:
                    print('Statistics Populated..')
            except IndexError:
                print('Index Error')
        print(template_to_df_map[entity])        
    Write_Normal_Output(df_Config,template_to_df_map,opco)
    Archive_File(df_Config,Input_file)


# In[10]:


def Write_Normal_Output(df_Config,template_to_df_map,opco):
    #Create Output File to Fill to fill data
    #import shutil
    Output_File_Path = df_Config[df_Config['CODENAME']=='OUTPUT_FILE_PATH']['CODEVALUE']
    Output_Temp_File_Path = df_Config[df_Config['CODENAME']=='OUTPUT_TEMPLATE_PATH']['CODEVALUE'].iloc[0]
    Output_File_Name = df_Config[df_Config['CODENAME']=='OUTPUT_FILE_NAME']['CODEVALUE'].iloc[0]
    Output_File = glob.glob(Output_Temp_File_Path+Output_File_Name)
    Output_File = shutil.copy2(Output_File[0], Output_File_Path.iloc[0]+"MTN-Backup_Tracker_OUT_"+str(datetime.date.today())+".xlsx") # complete target filename given
    CopyToFile = Output_File_Path.iloc[0]+"MTN-Backup_Tracker_OUT_"+str(datetime.date.today())+".xlsx"
    Output_File = Output_File_Path.iloc[0]+"MTN-Backup_Tracker_OUT_"+str(datetime.date.today())+".xlsx"
    print(Output_File)

    book = openpyxl.load_workbook(Output_File)
    writer = pd.ExcelWriter(Output_File, engine='openpyxl') 
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    #df_Day = pd.DataFrame({'Day': [pd.to_datetime('today').strftime('%m-%d-%Y')]})
    df_Day = pd.DataFrame({'Day': [datetime.datetime.today().strftime("%d-%b-%Y")]})
    df_Day_mdy = pd.DataFrame({'Day': [datetime.datetime.today().strftime("%m/%d/%Y")]})
    

    for i,row_output_page in df_Config.loc[df_Config['CODENAME']=='ENTITY_OUTPUT_PAGE',['CODEVALUE','CODEVALUE1','CODEVALUE2']].iterrows():
        #print(row_output_page)
        if row_output_page['CODEVALUE'] not in ['Summary','Dashboard_Backup_Success_Fail','Dashboard_Performance']:
            template_to_df_map[row_output_page['CODEVALUE']].to_excel(writer, row_output_page['CODEVALUE1'], index=False,startrow=row_output_page['CODEVALUE2'], header=False)
            if opco == 'NG':
                for i,row_stat in df_Config.loc[df_Config['CODENAME']=='STAT_'+row_output_page['CODEVALUE'],['CODEVALUE','CODEVALUE1','CODEVALUE2']].iterrows():
                    Is_Exists,Stat = Find_Stat(template_to_df_map['Summary'].iloc[0:5,:],row_output_page['CODEVALUE'])
                    if Is_Exists =='Y':
                        #print(Stat.iloc[:,-3])
                        Stat.to_excel(writer, row_output_page['CODEVALUE1'], index=False,startrow=row_stat['CODEVALUE1'],startcol=row_stat['CODEVALUE2'], header=True)
                        df_Day.to_excel(writer, row_output_page['CODEVALUE1'], index=False,startrow=row_stat['CODEVALUE1']-1,startcol=row_stat['CODEVALUE2'], header=False)          
        
        elif opco == 'NG' and row_output_page['CODEVALUE'] =='Summary':
            #print(template_to_df_map[row_output_page['CODEVALUE']].iloc[0:5,:])
            template_to_df_map[row_output_page['CODEVALUE']].iloc[0:5,:].to_excel(writer, row_output_page['CODEVALUE1'], index=False,startrow=row_output_page['CODEVALUE2'], header=False)
            df_Day_mdy.to_excel(writer, row_output_page['CODEVALUE1'], index=False,startrow=1,startcol=4, header=False)
            print('Statistics populated..')
        elif opco == 'NG' and row_output_page['CODEVALUE'] =='Dashboard_Backup_Success_Fail':
            #df_stat['Day'] = pd.Series(pd.to_datetime('today').strftime('%d/%m/%Y'), index=df_stat.index)
            df_Day.to_excel(writer, row_output_page['CODEVALUE1'], index=False,startrow=int(datetime.datetime.today().strftime("%d")),startcol=0, header=False)
            df_row_stat = template_to_df_map['Summary'].loc[template_to_df_map['Summary']['Node Type']=='Total',['Backup Successful','Backup failed']]
            #print(template_to_df_map['Summary'].iloc[5:6,:]['Total number of backups','Backup failed','Backup Successful'])
            df_row_stat.to_excel(writer, row_output_page['CODEVALUE1'], index=False,startrow=int(datetime.datetime.today().strftime("%d")),startcol=1, header=False)
        elif opco == 'NG' and row_output_page['CODEVALUE'] =='Dashboard_Performance':
            #df_stat['Day'] = pd.Series(pd.to_datetime('today').strftime('%d/%m/%Y'), index=df_stat.index)
            df_Day.to_excel(writer, row_output_page['CODEVALUE1'], index=False,startrow=int(datetime.datetime.today().strftime("%d")),startcol=0, header=False)
            df_row_stat = template_to_df_map['Summary'].loc[template_to_df_map['Summary']['Node Type']=='Performance',['Backup Successful']]
            #print(template_to_df_map['Summary'].iloc[5:6,:]['Total number of backups','Backup failed','Backup Successful'])
            df_row_stat.to_excel(writer, row_output_page['CODEVALUE1'], index=False,startrow=int(datetime.datetime.today().strftime("%d")),startcol=1, header=False)
            

        
    writer.save()
    if opco == 'NG':
        Prepare_Chart(Output_File,opco)
        print('Graph Plotted...')
    Send_Email_SMTP(Output_File_Path.iloc[0],"MTN-Backup_Tracker_OUT_"+str(datetime.date.today())+".xlsx",opco,df_Config)
    #Send_Email(CopyToFile,opco)
    Archive_File(df_Config,CopyToFile)


# In[11]:


def Prepare_Chart(Output_File,opco):

    if opco == 'NG':
        wb = openpyxl.load_workbook(Output_File)
        ws = wb['Dashboard_Backup_Success_Fail']
        ws1 =wb['Dashboard_Performance']
        data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=32)
        # Chart with date axis
        c1 = LineChart()
        c1.height = 10 # default is 7.5
        c1.width = 40 # default is 15
        c1.title = "BACKUP SUCCESS"
        c1.style = 12
        #c2.y_axis.title = "Failure"
        c1.y_axis.crossAx = 500
        c1.x_axis = DateAxis(crossAx=100)
        c1.x_axis.number_format = 'd-mmm'
        c1.x_axis.majorTimeUnit = "days"
        #c2.x_axis.title = "Date"
        #s1 = c2.series[0]
        #s1.smooth = True # Make the line smooth

        c1.add_data(data, titles_from_data=True)
        s1 = c1.series[0]
        s1.graphicalProperties.line.solidFill = "00AAAA"
        #c2.add_data(data, titles_from_data=False)
        dates = Reference(ws, min_col=1, min_row=2, max_row=32)
        c1.set_categories(dates)

        ws.add_chart(c1, "F19")
        ###################
        data = Reference(ws, min_col=3, min_row=1, max_col=3, max_row=32)
        c2 = LineChart()
        c2.height = 10 # default is 7.5
        c2.width = 40 # default is 15
        c2.title = "BACKUP FAILED"
        c2.style = 12
        #c2.y_axis.title = "Failure"
        c2.y_axis.crossAx = 500
        c2.x_axis = DateAxis(crossAx=100)
        c2.x_axis.number_format = 'd-mmm'
        c2.x_axis.majorTimeUnit = "days"
        #c2.x_axis.title = "Date"
        #s1 = c2.series[0]
        #s1.smooth = True # Make the line smooth

        c2.add_data(data, titles_from_data=True)
        #c2.add_data(data, titles_from_data=False)
        dates = Reference(ws, min_col=1, min_row=2, max_row=32)
        c2.set_categories(dates)

        ws.add_chart(c2, "F1")
        ##########################
        data = Reference(ws1, min_col=2, min_row=1, max_col=2, max_row=32)
        # Chart with date axis
        c3 = LineChart()
        c3.height = 10 # default is 7.5
        c3.width = 40 # default is 15
        c3.title = "PERFORMANCE"
        c3.style = 12
        #c2.y_axis.title = "Failure"
        c3.y_axis.crossAx = 500
        c3.x_axis = DateAxis(crossAx=100)
        c3.x_axis.number_format = 'd-mmm'
        c3.x_axis.majorTimeUnit = "days"
        #c2.x_axis.title = "Date"
        #s1 = c2.series[0]
        #s1.smooth = True # Make the line smooth

        c3.add_data(data, titles_from_data=True)
        s3 = c3.series[0]
        s3.graphicalProperties.line.solidFill = "00AAAA"
        #c2.add_data(data, titles_from_data=False)
        dates = Reference(ws1, min_col=1, min_row=2, max_row=32)
        c3.set_categories(dates)

        ws1.add_chart(c3, "F1")
        #########################

        wb.save(Output_File)
        print('Done..')


# In[12]:


def Archive_File(df_Config,Input_file):
        
    #Archive Input
    Archive_File_Path = df_Config[df_Config['CODENAME']=='ARCHIVE_FILE_PATH']['CODEVALUE'].iloc[0]
    Archive_File = glob.glob(Archive_File_Path)
    Archive_File = shutil.move(Input_file, Archive_File[0]) # complete target filename given


# In[13]:


# def Read_Email(opco,df_Config):
    # Input_File_Path = df_Config[df_Config['CODENAME']=='INPUT_FILE_PATH']['CODEVALUE']
    # outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    # accounts= win32com.client.Dispatch("Outlook.Application").Session.Accounts
    # for account in accounts:
        # if account.DisplayName !='Outlook2010':
            # continue
        # global inbox
        # inbox = outlook.Folders(account.DeliveryStore.DisplayName)
        # #inbox = inbox.GetDefaultFolder(6).Folders.Item("IvoryCoast")
        # print("****Account Name**********************************")
        # #print(account.DisplayName,file=f)
        # print(account.DisplayName)
        # print("***************************************************")
        # #folders = inbox.Folders

        # root_folder = inbox.Folders(2).Folders.Item(GetOpco_Name(opco))
        # print(root_folder)
        # messages = root_folder.Items
        # for message in messages:
            # if message.Unread == True:
                # print("New Mail Found... Downloading Attachment...")
                # #Loop to check if the attachment name is the same
                # for attachments in message.Attachments:
                    # if 'MTN-Nigeria-Backup-Tracker' in attachments.Filename:
                        # attachment_name = attachments.Filename
                    # #if attachments.Filename == attachment_name:
                        # #Saves to the attachment to the working directory 
                        # #attachments.SaveASFile(os.getcwd() + '\\' + 'my_attachment_name' + date_time_stamp + '.xlsx')
                        # attachments.SaveASFile(Input_File_Path[0]  + attachment_name)
                        # print (attachments)
                        # time.sleep(2)
                        # message.Unread = False
                        # break
                    # #Go to next unread messages if any
                # message = messages.GetNext()
        # else:
                # print ("Checking...")



# In[14]:


# def Send_Email(Attachment_Path,opco,df_Config):
    # s = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    # o = win32com.client.Dispatch("Outlook.Application")
    # s.Logon("Outlook2010")
    # Msg = o.CreateItem(0)
    # Msg.To = "rajarshi.roy@ericsson.com"
    
    # Msg.CC = "rajarshi.roy@ericsson.com"
    # #Msg.BCC = "more email addresses here"

    # Msg.Subject = "MTN Backup Tracker:"+ GetOpco_Name(opco)
    # Msg.Body = "This is MTN backup tracker"
    # attachment1 = Attachment_Path
    # #attachment1 = "C:\\Users\\erairoy\\Desktop\\Automation\\MTN\\OUTPUT\\IV\\"+'MTN-Backup_Tracker_OUT_2019-03-15.xlsx'
    # #attachment2 = "Path to attachment no. 2"
    # Msg.Attachments.Add(attachment1)
    # #Msg.Attachments.Add(attachment2)

    # Msg.Send()
    # print('Email Sent..')


# In[15]:


def Send_Email_SMTP(Attachment_Path,Attachment_Name,opco,df_Config):
	subject = 'MTN Backup Tracker for '+ GetOpco_Name(opco)
	text = 'MTN Backup Tracker for '+ GetOpco_Name(opco)
	fromaddr = 'no-reply@AutoBOT' #df_Config[df_Config['CODENAME']=='EMAIL_TO_SEND']['CODEVALUE'].iloc[0]
	toaddr = ['gnoc.1st.la.mtn.rsaa@ericsson.com','PDLMTNFMIN@pdl.internal.ericsson.com','PDLFOINMTN@pdl.internal.ericsson.com','PDLHIFTMAN@pdl.internal.ericsson.com']#df_Config[df_Config['CODENAME']=='EMAIL_TO_SEND']['CODEVALUE'].iloc[0]
	COMMASPACE = ', '
	#toaddr = [df_Config[df_Config['CODENAME']=='EMAIL_TO_SEND']['CODEVALUE'].iloc[0],df_Config[df_Config['CODENAME']=='EMAIL_TO_SEND']['CODEVALUE1'].iloc[0]]
	

	msg = MIMEMultipart()

	msg['From'] = fromaddr
	msg['To'] = COMMASPACE.join(toaddr)
	msg['Subject'] = subject

	body = text

	msg.attach(MIMEText(body, 'plain'))

	filename = Attachment_Name
	#attachment = open('/home/congo/PM-Automation/MTN/ARCHIVE/IV/')#open("PATH OF THE FILE", "rb")

	part = MIMEBase('application', 'octet-stream')
	#part.set_payload((attachment).read())
	part.set_payload(open(Attachment_Path+Attachment_Name, "rb").read())

	encoders.encode_base64(part)
	part.add_header('Content-Disposition', "attachment; filename= %s" % filename)

	msg.attach(part)
	text = msg.as_string()

	try:
	   smtpObj = smtplib.SMTP('172.23.168.13',25)
	   smtpObj.sendmail(fromaddr, toaddr, text)
	   print("Successfully sent email")
	except smtplib.SMTPException:
	   print("Error: unable to send email")


# In[16]:


def GetOpco_Name(opco):
    if opco =='NG':
        return 'Nigeria'
    elif opco == 'IV':
        return 'IvoryCoast'
    elif opco == 'SW':
        return 'Swaziland'
    elif opco == 'BIS':
        return 'Bissau'
    elif opco == 'CONG':
        return 'Congo'
    elif opco == 'BN':
        return 'Benin'
    elif opco == 'GH':
        return 'Ghana'
    


# In[17]:


if __name__ == '__main__':
    print(len(sys.argv))
    if len(sys.argv) < 2:
        print('To few arguments, please specify a filename')
    #filename = sys.argv[1]
    #print('Filename:', filename)
    #print(sys.argv[0])
    #print(sys.argv[2])
        
    for i,opco in Configsheet_to_df_map['OPCO'].iterrows():
        Load_Format_Write(opco.iloc[0],Configsheet_to_df_map)
        print(opco.iloc[0])

