
import sys

reload(sys)

sys.setdefaultencoding('utf8')

import logging
from datetime import date
import datetime

import openpyxl
import pandas as pd
from utils import *
import performence

print("Process Started")
day = datetime.datetime.today() 
day_DD_MON_YYYY = day.strftime("%d-%b-%Y")
# Initialising Logging module
log_file_path = "./logs/logs_{}.log".format(str(date.today()))
logging.basicConfig(
    level=logging.INFO,
    # filename=log_file_path
)

logging.info("Starting Automation...")

# Loading config.ini file into a dictionary
logging.info("Loading Config.ini")
config_data = load_config_ini()

# initialising variable from config.ini
logging.info("Loading Variables from config.ini ")
Base_path = config_data['base_dir_inp_files']
Template_file = config_data['template_file']
Output_File = config_data['output_file_path'] + \
              config_data['output_file_prefix'] + \
              str(date.today()) + ".xlsx"
pma_nw_file_path = config_data['pma_nw_file_path']


try:

    # reading pma_nw_final.conf and creating dictionary
    logging.info("Loading Configurations from pma_nw_file_path.conf")
    updated_pma_dict = pma_nw_node_dict(pma_nw_file_path)

    # Creating File which we will have in Output
    logging.info("Creating Object for output Excel ")
    writer = pd.ExcelWriter(Output_File, engine='openpyxl')
    book = openpyxl.load_workbook(Template_file)
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    # Loading template in pandas dataframe
    logging.info("Reading Input Template file")
    template_df_map = pd.read_excel(Template_file, sheet_name=None)

    # Generating Dictionary after reading and parsing node types directories
    # based on nedate and its corresponding Inp file
    logging.info("Creating dictionary of all node types after parsing ini files")
    parsed_hash = main()
    overall_nodes = 0
    overall_success_nodes = 0
    summary_data = []
    # Iterating each node type for Update individual output sheet
    for key, values in parsed_hash.items():
        opco = key
        print(opco)

        # Continue if opco is commented in pma_nw-final.cong
        if opco.lower() not in updated_pma_dict.keys():
            continue
        
        # # ------------------
        if 'SDP' == opco.upper():
            sheet_name = "SDP_Daily"
            final_data, success_nodes, success_nodes_without_pair = nodes_iterator_sdp(
                values, updated_pma_dict[opco.lower()]
                )
            total_nodes = len(final_data)
            writer.sheets[sheet_name].cell(1,4).value  = day_DD_MON_YYYY
            writer.sheets[sheet_name].cell(2,4).value  = total_nodes
            writer.sheets[sheet_name].cell(3,4).value  = total_nodes-success_nodes
            writer.sheets[sheet_name].cell(4,4).value  = success_nodes
            final_data = pd.DataFrame (final_data)
            template_df_map[sheet_name] = final_data
            template_df_map[sheet_name].to_excel(
                writer, sheet_name, index=False, startrow=6,startcol=0, header=False
                )
            summary = ['SDP','Tape Backup',total_nodes*2,success_nodes_without_pair, total_nodes*2-success_nodes_without_pair ]
            summary_data.append(summary)
            overall_nodes = overall_nodes+total_nodes*2
            overall_success_nodes = overall_success_nodes+success_nodes_without_pair
            

            # SDP geo-----------------
            values  = parsed_hash["sdp-geo-red"]
            sheet_name = "SDP_Geo_Redundancy_Status"
            final_data, success_nodes = nodes_iterator_sdp_geo(
                values, updated_pma_dict[opco.lower()]
                )
            total_nodes = len(final_data)
            writer.sheets[sheet_name].cell(1,3).value  = day_DD_MON_YYYY

            final_data = pd.DataFrame (final_data)
            template_df_map[sheet_name] = final_data
            template_df_map[sheet_name].to_excel(
                writer, sheet_name, index=False, startrow=1,startcol=0, header=False
                )
            overall_nodes = overall_nodes+total_nodes
            overall_success_nodes = overall_success_nodes+success_nodes


        # # ------------------
        if 'AIR' == opco.upper():
            sheet_name = "AIR_Daily"
            final_data, success_nodes = nodes_iterator_air(
                values, updated_pma_dict[opco.lower()]
                )
            total_nodes = len(final_data)
            writer.sheets[sheet_name].cell(1,3).value  = day_DD_MON_YYYY
            writer.sheets[sheet_name].cell(2,3).value  = total_nodes
            writer.sheets[sheet_name].cell(3,3).value  = total_nodes-success_nodes
            writer.sheets[sheet_name].cell(4,3).value  = success_nodes
            final_data = pd.DataFrame (final_data)
            template_df_map[sheet_name] = final_data
            template_df_map[sheet_name].to_excel(
                writer, sheet_name, index=False, startrow=6,startcol=0, header=False
                )
            summary = ['AIR','Tape Backup',total_nodes,success_nodes, total_nodes-success_nodes ]
            summary_data.append(summary)
            overall_nodes = overall_nodes+total_nodes
            overall_success_nodes = overall_success_nodes+success_nodes

        # # ------------------
        if 'CCN' == opco.upper():
            sheet_name = "CCN_Daily"
            final_data, success_nodes = nodes_iterator_ccn(
                values, updated_pma_dict[opco.lower()]
                )
            total_nodes = len(final_data)
            writer.sheets[sheet_name].cell(1,6).value  = day_DD_MON_YYYY
            writer.sheets[sheet_name].cell(2,6).value  = total_nodes
            writer.sheets[sheet_name].cell(3,6).value  = total_nodes-success_nodes
            writer.sheets[sheet_name].cell(4,6).value  = success_nodes
            final_data = pd.DataFrame (final_data)
            template_df_map[sheet_name] = final_data
            template_df_map[sheet_name].to_excel(
                writer, sheet_name, index=False, startrow=6,startcol=0, header=False
                )
            summary = ['CCN','DBN Backup',total_nodes,success_nodes, total_nodes-success_nodes ]
            summary_data.append(summary)
            overall_nodes = overall_nodes+total_nodes
            overall_success_nodes = overall_success_nodes+success_nodes

        # ------------------
        if 'NGVS' == opco.upper():
            sheet_name = "NGVS_DAILY"
            final_data, success_nodes_cassendra, success_nodes_tape = nodes_iterator_ngvs(
                values, updated_pma_dict[opco.lower()]
                )
            total_nodes = len(final_data)
            writer.sheets[sheet_name].cell(1,3).value  = day_DD_MON_YYYY
            writer.sheets[sheet_name].cell(3,3).value  = total_nodes
            writer.sheets[sheet_name].cell(4,3).value  = total_nodes-success_nodes_cassendra
            writer.sheets[sheet_name].cell(5,3).value  = success_nodes_cassendra
            writer.sheets[sheet_name].cell(3,4).value  = total_nodes
            writer.sheets[sheet_name].cell(4,4).value  = total_nodes-success_nodes_tape
            writer.sheets[sheet_name].cell(5,4).value  = success_nodes_tape
            final_data = pd.DataFrame (final_data)
            template_df_map[sheet_name] = final_data
            template_df_map[sheet_name].to_excel(
                writer, sheet_name, index=False, startrow=5,startcol=0, header=False
                )
            summary = ['NGVS','FS backup/DB Backup',total_nodes,success_nodes, total_nodes-success_nodes ]
            summary_data.append(summary)
            overall_nodes = overall_nodes+total_nodes
            overall_success_nodes = overall_success_nodes+success_nodes


            # NGVS geo-----------------
            values  = parsed_hash["ngvs-geo-red"]
            sheet_name = "NGVS_Geo Redundancy_status"
            final_data, success_nodes = nodes_iterator_ngvs_geo(
                values, updated_pma_dict[opco.lower()]
                )
            total_nodes = len(final_data)
            writer.sheets[sheet_name].cell(1,3).value  = day_DD_MON_YYYY

            final_data = pd.DataFrame (final_data)
            template_df_map[sheet_name] = final_data
            template_df_map[sheet_name].to_excel(
                writer, sheet_name, index=False, startrow=1,startcol=0, header=False
                )
            overall_nodes = overall_nodes+total_nodes
            overall_success_nodes = overall_success_nodes+success_nodes

        # # ------------------
        if 'OCC' == opco.upper():
            sheet_name = "OCC_Daily"
            final_data, success_nodes = nodes_iterator_occ(
                values, updated_pma_dict[opco.lower()]
                )
            total_nodes = len(final_data)
            writer.sheets[sheet_name].cell(1,5).value  = day_DD_MON_YYYY
            writer.sheets[sheet_name].cell(2,5).value  = total_nodes
            writer.sheets[sheet_name].cell(3,5).value  = total_nodes-success_nodes
            writer.sheets[sheet_name].cell(4,5).value  = success_nodes

            final_data = pd.DataFrame (final_data)
            template_df_map[sheet_name] = final_data
            template_df_map[sheet_name].to_excel(
                writer, sheet_name, index=False, startrow=5,startcol=0, header=False
                )
            summary = ['OCC','FS Backup',total_nodes,success_nodes, total_nodes-success_nodes ]
            summary_data.append(summary)
            overall_nodes = overall_nodes+total_nodes
            overall_success_nodes = overall_success_nodes+success_nodes


        # ------------------
        if 'VCCN' == opco.upper():
            sheet_name = "VCCN_Daily"
            final_data, success_nodes = nodes_iterator_vccn(
                values, updated_pma_dict[opco.lower()]
                )
            total_nodes = len(final_data)
            writer.sheets[sheet_name].cell(1,3).value  = day_DD_MON_YYYY
            writer.sheets[sheet_name].cell(2,3).value  = total_nodes
            writer.sheets[sheet_name].cell(3,3).value  = total_nodes-success_nodes
            writer.sheets[sheet_name].cell(4,3).value  = success_nodes

            final_data = pd.DataFrame (final_data)
            template_df_map[sheet_name] = final_data
            template_df_map[sheet_name].to_excel(
                writer, sheet_name, index=False, startrow=6,startcol=0, header=False
                )
            summary = ['VCCN','VCCN Backup',total_nodes,success_nodes, total_nodes-success_nodes ]
            summary_data.append(summary)
            overall_nodes = overall_nodes+total_nodes
            overall_success_nodes = overall_success_nodes+success_nodes

    #  Writting data in summary sheet
    sheet_name = "Summary"
    total_nodes = len(summary_data)
    writer.sheets[sheet_name].cell(1,5).value  = day_DD_MON_YYYY

    summary_data = pd.DataFrame (summary_data)
    template_df_map[sheet_name] = summary_data
    template_df_map[sheet_name].to_excel(
        writer, sheet_name, index=False, startrow=2,startcol=0, header=False
        )

 
    # Flattening Json and return a list of nodes of NotDone cases for Logs sheet.
    logging.info("Flatting Dataframe")
    log_rows = json_flatten_for_logs(parsed_hash, updated_pma_dict, Base_path)

    logging.info("Writting Data to Logs sheet in output excel file")
    if len(log_rows) > 0:
        template_df_map['Logs'] = template_df_map['Logs'] \
            .append(log_rows, ignore_index=True)

        template_df_map['Logs'].to_excel(
            writer, 'Logs', index=False, startrow=1, startcol=0, header=False
        )

    print("===================================")
    # print(summary_data)
    print(template_df_map["Logs"])
    print("===================================")




    performence_per = overall_success_nodes/float(overall_nodes)
    performence.create_performence_csv(overall_nodes, overall_success_nodes, performence_per)


    # Preparing Dashboard for Performence ==============
    sheet_name = "Dashboard_Performance"
    final_data = performence.read_performence_csv()
    total_nodes = len(final_data)
    final_data = pd.DataFrame (final_data)
    template_df_map[sheet_name] = final_data
    template_df_map[sheet_name].to_excel(
        writer, sheet_name, index=False, startrow=1,startcol=0, header=False
        )
    # Preparing Dashboard for success and fail =========
    sheet_name = "Dashboard_Backup_Success_Fail"
    final_data = performence.read_success_fail_csv()
    total_nodes = len(final_data)
    final_data = pd.DataFrame (final_data)
    template_df_map[sheet_name] = final_data
    template_df_map[sheet_name].to_excel(
        writer, sheet_name, index=False, startrow=1,startcol=0, header=False
        )








    writer.save()
    logging.info("Sending mail")
    flag=False
    # Send_Email_SMTP(Output_File,flag)
except Exception as error:
    raise
    logging.error("Error Occured: {}".format(error))
    logging.info("Sending Failier mail")
    flag=True
    # Send_Email_SMTP(log_file_path,flag)
logging.info("Automation Ended")
print("Process Ended")

